from django.shortcuts import get_object_or_404
from django.contrib.auth.decorators import login_required
import openpyxl
from openpyxl.styles import Font, Alignment
from django.http import HttpResponse
from .models import Category, Product,Sale
from datetime import datetime,timedelta
from django.db.models import F
from django.db.models import Sum


@login_required
def export_products_excel(request):
    category_id = request.GET.get('category_id')

    # Create a new workbook and select the active sheet
    wb = openpyxl.Workbook()
    ws = wb.active

    # Set up the header row
    headers = ['Category', 'Brand', 'Product Name', 'Stock', 'Purchase Price', 'Sale Price']
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')

    # Fetch the data
    if category_id:
        category = get_object_or_404(Category, id=category_id)
        products = Product.objects.filter(brand__category=category).select_related('brand__category', 'brand__brand_name')
        filename = f"{category.category_name}_products.xlsx"
    else:
        products = Product.objects.all().select_related('brand__category', 'brand__brand_name')
        filename = "all_products.xlsx"

    # Populate the sheet with data
    for row, product in enumerate(products, start=2):
        ws.cell(row=row, column=1, value=product.brand.category.category_name)
        ws.cell(row=row, column=2, value=product.brand.brand_name.brand_name)
        ws.cell(row=row, column=3, value=product.product_name)
        ws.cell(row=row, column=4, value=product.quantity_in_stock)
        ws.cell(row=row, column=5, value=product.purchase_price)
        ws.cell(row=row, column=6, value=product.sale_price)

    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(cell.value)
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[column_letter].width = adjusted_width

    # Create the HttpResponse object with Excel mime type
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    # Save the workbook to the response
    wb.save(response)

    return response

from openpyxl.styles import Font, Alignment, PatternFill
@login_required
def export_sales_excel(request):
    # Get all sales for the logged-in employee
    all_sales = Sale.objects.filter(employee=request.user).order_by('-sale_date')

    # Get filter dates from request
    start_date = request.GET.get('start_date')
    end_date = request.GET.get('end_date')

    # Apply date filters if provided
    if start_date:
        all_sales = all_sales.filter(sale_date__gte=datetime.strptime(start_date, '%Y-%m-%d'))
    if end_date:
        all_sales = all_sales.filter(sale_date__lte=datetime.strptime(end_date, '%Y-%m-%d') + timedelta(days=1))

    # Calculate filtered total money made
    filtered_total = all_sales.aggregate(
        total=Sum(F('quantity_sold') * F('product__sale_price'))
    )['total'] or 0

    # Create a new workbook and select the active sheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sales History"

    # Define styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    centered_alignment = Alignment(horizontal='center')

    # Set up the header row
    headers = ['Date', 'Product', 'Category', 'Brand', 'Quantity', 'Unit Price', 'Total Price']
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = centered_alignment

    # Populate the sheet with data
    for row, sale in enumerate(all_sales, start=2):
        ws.cell(row=row, column=1, value=sale.sale_date.strftime("%Y-%m-%d %H:%M"))
        ws.cell(row=row, column=2, value=sale.product.product_name)
        ws.cell(row=row, column=3, value=sale.product.brand.category.category_name)
        ws.cell(row=row, column=4, value=sale.product.brand.brand_name.brand_name)  
        ws.cell(row=row, column=5, value=sale.quantity_sold)
        ws.cell(row=row, column=6, value=float(sale.product.sale_price))
        ws.cell(row=row, column=7, value=float(sale.quantity_sold * sale.product.sale_price))

    # Add total row
    total_row = len(all_sales) + 2
    ws.cell(row=total_row, column=6, value="Total Sales:")
    ws.cell(row=total_row, column=7, value=float(filtered_total))

    # Style the total row
    for col in range(6, 8):
        cell = ws.cell(row=total_row, column=col)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="E6E6E6", end_color="E6E6E6", fill_type="solid")

    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[column_letter].width = adjusted_width

    # Create the HttpResponse object with Excel mime type
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="sales_history.xlsx"'

    # Save the workbook to the response
    wb.save(response)

    return response